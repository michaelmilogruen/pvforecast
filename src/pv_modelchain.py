"""
PV System Simulation using pvlib's ModelChain

This script implements a PV system simulation using pvlib's high-level ModelChain interface.
It calculates the power output of a photovoltaic system based on provided parameters and
compars it with the manual approach used in pvsimmodule.py.

Author: Michael Grün
Email: michaelgruen@hotmail.com
Version: 1.2 (Modified to keep plots open)
Date: 2025-05-07
"""

import os
from typing import Tuple, Optional, Dict, Union

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pvlib
from pvlib.location import Location
from pvlib.pvsystem import PVSystem, Array, FixedMount
from pvlib.modelchain import ModelChain
from pvlib.temperature import TEMPERATURE_MODEL_PARAMETERS
import poadata
from datetime import datetime
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy import stats


# Define parameters
PARAMS = {
    'start': '2022-01-01 01:10',
    'end': '2024-01-01 00:10',
    'latitude': 47.3877,
    'longitude': 15.0941,
    'tilt': 30,
    'azimuth': 149.716,
    'celltype': 'monocrystalline',
    'pdc0': 240,
    'v_mp': 29.87,
    'i_mp': 8.04,
    'v_oc': 37.33,
    'i_sc': 8.78,
    'alpha_sc': 0.0041,
    'beta_voc': -0.114,
    'gamma_pdc': -0.405,
    'cells_in_series': 69,
    'temp_ref': 25,
    # PVWatts system losses parameters (in percentage) - lowered values
    'losses_parameters': {
        'soiling': 1.5,
        'shading': 0.5,
        'snow': 0.2,
        'mismatch': 1.0,
        'wiring': 0.8,
        'connections': 0.3,
        'nameplate_rating': 0.7,
        'age': 1.0,
        'availability': 0.5
    }
}


def execute_vba_actions(file_name: str) -> None:
    """
    Format Excel file and copy columns between worksheets.

    Args:
        file_name (str): The name of the Excel file to be processed.

    Returns:
        None
    """
    workbook = load_workbook(file_name)
    model_chain_results = workbook['Model Chain Results']
    poa_data = workbook['POA Data']

    # Autofit all columns in both worksheets
    for worksheet in [model_chain_results, poa_data]:
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
            adjusted_width = (max_length + 2) * 1.2
            column_letter = get_column_letter(column[0].column)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Copy columns B:J from 'POA Data' to 'Model Chain Results' starting at column H
    for col_idx, source_col in enumerate(range(2, 11), start=8):  # B:J -> H onwards
        for row_idx, cell in enumerate(poa_data.iter_rows(min_row=1, min_col=source_col, max_col=source_col), start=1):
            model_chain_results.cell(row=row_idx, column=col_idx, value=cell[0].value)

    # Save the modified workbook
    workbook.save(file_name)


def create_cec_system(params: dict) -> PVSystem:
    """
    Create a PVSystem object configured with CEC model parameters.

    Args:
        params (dict): Parameters for the PV system

    Returns:
        PVSystem: Configured PV system object
    """
    # Fit the CEC single diode model parameters from module characteristics
    I_L_ref, I_o_ref, R_s, R_sh_ref, a_ref, Adjust = pvlib.ivtools.sdm.fit_cec_sam(
        celltype=params['celltype'],
        v_mp=params['v_mp'],
        i_mp=params['i_mp'],
        v_oc=params['v_oc'],
        i_sc=params['i_sc'],
        alpha_sc=params['alpha_sc'],
        beta_voc=params['beta_voc'],
        gamma_pmp=params['gamma_pdc'],
        cells_in_series=params['cells_in_series']
    )

    # Configure module parameters dictionary
    module_parameters = {
        'alpha_sc': params['alpha_sc'],
        'a_ref': a_ref,
        'I_L_ref': I_L_ref,
        'I_o_ref': I_o_ref,
        'R_sh_ref': R_sh_ref,
        'R_s': R_s,
        'Adjust': Adjust
    }

    # Access temperature model parameters using the correct nested structure
    # 'sapm' is the temperature model, 'open_rack_glass_polymer' is the configuration
    temperature_model_parameters = TEMPERATURE_MODEL_PARAMETERS['sapm']['open_rack_glass_polymer']

    # Create array with fixed mount at specified tilt and azimuth
    array = Array(
        mount=FixedMount(surface_tilt=params['tilt'], surface_azimuth=params['azimuth']),
        module_parameters=module_parameters,
        temperature_model_parameters=temperature_model_parameters,
        modules_per_string=14,
        strings=3
    )

    # Retrieve the Fronius inverter parameters
    cec_inverters = pvlib.pvsystem.retrieve_sam('CECInverter')
    inverter = cec_inverters['Fronius_International_GmbH__Fronius_Symo_12_5_3_480__480V_']

    # Create and return the PVSystem
    return PVSystem(arrays=[array], inverter_parameters=inverter)


def calculate_system_losses(params: dict) -> Dict[str, float]:
    """
    Calculate system losses parameters based on the input parameters.

    Args:
        params (dict): Dictionary containing loss parameters

    Returns:
        Dict[str, float]: Dictionary of loss parameters for the pvlib losses model
    """
    # Calculate total DC losses
    dc_loss_factors = [
        params.get('soiling', 0.03),           # Soiling losses
        params.get('mismatch', 0.02),          # Module mismatch
        params.get('wiring', 0.015),           # DC wiring losses
        params.get('connections', 0.005),      # Connection losses
        params.get('lid', 0.01),               # Light-induced degradation
        params.get('nameplate_rating', 0.015), # Nameplate rating loss
        params.get('age', 0.02),               # Age-related degradation
        params.get('availability', 0.01)       # System availability/downtime
    ]

    # Convert from individual loss factors to derate factors
    dc_derate_factors = [1 - loss for loss in dc_loss_factors]

    # Calculate total DC derate factor (multiplicative)
    total_dc_derate = np.prod(dc_derate_factors)

    # Calculate total DC loss
    total_dc_loss = 1 - total_dc_derate

    return {
        'dc_ohmic_percent': 100 * params.get('wiring', 0.015),  # DC ohmic losses in percent
        'ac_ohmic_percent': 0.5,                               # AC ohmic losses in percent
        'soiling_percent': 100 * params.get('soiling', 0.03),  # Soiling losses in percent
        'total_irrad_loss_percent': 100 * total_dc_loss        # Combined losses in percent
    }


def calculate_power_modelchain(params: dict) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Calculate PV power output using pvlib's ModelChain.

    Args:
        params (dict): Parameters for the PV system and simulation

    Returns:
        Tuple[pd.DataFrame, pd.DataFrame]: AC power results and POA data
    """
    # Create location object
    location = Location(
        latitude=params['latitude'],
        longitude=params['longitude'],
        tz='Europe/Vienna',
        altitude=547.6,
        name='EVT'
    )

    # Get POA irradiance data (same as in original script)
    poa_data = poadata.get_pvgis_data(
        params['latitude'],
        params['longitude'],
        2022, 2023,
        params['tilt'],
        params['azimuth']
    )
    poa_data.to_csv("poa_data_Leoben_EVT_io.csv")
    poa_data = pd.read_csv('poa_data_Leoben_EVT_io.csv', index_col=0)
    poa_data.index = pd.date_range(start='2022-01-01 01:10', periods=len(poa_data.index), freq="h")
    poa_data = poa_data[params['start']:params['end']]

    # Create PV system with CEC configuration
    system = create_cec_system(params)

    # Set PVWatts losses parameters from the params dictionary
    system.losses_parameters = params.get('losses_parameters', {})

    # Create ModelChain object with appropriate models
    mc = ModelChain(
        system,
        location,
        aoi_model="ashrae",
        spectral_model="no_loss",
        dc_model="cec",
        ac_model="sandia",
        temperature_model="sapm",
        losses_model="pvwatts"  # Use pvwatts losses model to account for system losses
    )

    # Prepare input data for the model
    weather_data = pd.DataFrame({
        'poa_global': poa_data['poa_global'],
        'poa_direct': poa_data['poa_direct'],
        'poa_diffuse': poa_data['poa_diffuse'],
        'temp_air': poa_data['temp_air'],
        'wind_speed': poa_data['wind_speed']
    }, index=poa_data.index)

    # Run the ModelChain simulation with POA data
    mc.run_model_from_poa(weather_data)

    # Extract results from ModelChain
    ac_results = mc.results.ac

    # Overwrite any negative AC power values to 0
    ac_results = ac_results.clip(lower=0)


    # Create a DataFrame with all the model chain results
    results_df = pd.DataFrame({
        'AC Power': ac_results, # Use the modified ac_results
        'DC scaled P_mp': mc.results.dc['p_mp'],
        'DC scaled V_mp': mc.results.dc['v_mp'],
        'DC scaled I_mp': mc.results.dc['i_mp'],
        'Cell Temperature': mc.results.cell_temperature
    })

    # Save results to Excel and CSV
    with pd.ExcelWriter("results_modelchain.xlsx") as writer:
        results_df.to_excel(writer, sheet_name='Model Chain Results')
        poa_data.to_excel(writer, sheet_name='POA Data')

        # Add a losses summary sheet
        losses_summary = pd.DataFrame({
            'Loss Type': [
                'Soiling',
                'Shading',
                'Snow',
                'Mismatch',
                'Wiring',
                'Connections',
                'Light-Induced Degradation',
                'Nameplate Rating',
                'Age Degradation',
                'Availability',
                'Total Losses'
            ],
            'Loss Percentage (%)': [
                params.get('losses_parameters', {}).get('soiling', 3.0),
                params.get('losses_parameters', {}).get('shading', 1.0),
                params.get('losses_parameters', {}).get('snow', 0.5),
                params.get('losses_parameters', {}).get('mismatch', 2.0),
                params.get('losses_parameters', {}).get('wiring', 1.5),
                params.get('losses_parameters', {}).get('connections', 0.5),
                params.get('losses_parameters', {}).get('lid', 1.0),
                params.get('losses_parameters', {}).get('nameplate_rating', 1.5),
                params.get('losses_parameters', {}).get('age', 2.0),
                params.get('losses_parameters', {}).get('availability', 1.0),
                pvlib.pvsystem.pvwatts_losses(**params.get('losses_parameters', {}))
            ]
        })
        losses_summary.to_excel(writer, sheet_name='System Losses')

    # Merge and save as CSV
    merged_df = pd.concat([results_df, poa_data], axis=1)
    merged_df.to_csv("merged_results_modelchain.csv")

    return ac_results, poa_data


def calculate_statistics(predicted: pd.Series, actual: pd.Series) -> Dict[str, float]:
    """
    Calculate various statistical metrics to compare predicted and actual values.

    Args:
        predicted (pd.Series): Predicted values
        actual (pd.Series): Actual observed values

    Returns:
        Dict[str, float]: Dictionary containing various statistical metrics
    """
    # Remove any NaN values
    mask = ~(np.isnan(predicted) | np.isnan(actual))
    predicted_clean = predicted[mask]
    actual_clean = actual[mask]

    if len(predicted_clean) == 0:
        return {
            'rmse': np.nan,
            'r2': np.nan,
            'mape': np.nan,
            'smape': np.nan
        }

    # Calculate RMSE (Root Mean Square Error)
    rmse = np.sqrt(np.mean((predicted_clean - actual_clean) ** 2))

    # Calculate R² (Coefficient of determination)
    r2 = stats.pearsonr(predicted_clean, actual_clean)[0] ** 2

    # Calculate MAPE (Mean Absolute Percentage Error)
    # Avoid division by zero by excluding zeros in actual values
    non_zero = actual_clean != 0
    mape = np.mean(np.abs((actual_clean[non_zero] - predicted_clean[non_zero]) / actual_clean[non_zero])) * 100

    # Calculate SMAPE (Symmetric Mean Absolute Percentage Error)
    # Avoid division by zero
    denominator = np.abs(actual_clean) + np.abs(predicted_clean)
    non_zero_denom = denominator != 0
    smape = np.mean(2.0 * np.abs(predicted_clean[non_zero_denom] - actual_clean[non_zero_denom]) /
                     denominator[non_zero_denom]) * 100

    return {
        'rmse': rmse,
        'r2': r2,
        'mape': mape,
        'smape': smape
    }


def plot_results(ac_results: pd.Series, poa_data: pd.DataFrame = None) -> Dict[str, float]:
    """
    Create high-quality plots of PV system performance.
    Same functionality as in the original script.

    Args:
        ac_results (pd.Series): The Series containing the AC power results.
        poa_data (pd.DataFrame, optional): The DataFrame containing POA irradiance data.

    Returns:
        Dict[str, float]: Dictionary containing statistical metrics for model performance
    """
    # Set publication-quality plot style
    try:
        plt.style.use('seaborn-v0_8-whitegrid')  # Updated style name
    except:
        plt.style.use('seaborn-whitegrid')  # Fallback for older matplotlib

    plt.rcParams.update({
        'font.family': 'serif',
        'font.size': 12,
        'axes.labelsize': 14,
        'axes.titlesize': 16,
        'xtick.labelsize': 12,
        'ytick.labelsize': 12,
        'legend.fontsize': 12,
        'figure.titlesize': 18,
        'figure.figsize': (10, 6),
        'figure.dpi': 300,
        'savefig.dpi': 300,
        'savefig.bbox': 'tight',
        'savefig.pad_inches': 0.1
    })

    # Create results directory if it doesn't exist
    os.makedirs('results', exist_ok=True)

    # 1. Annual time series
    fig, ax = plt.subplots()
    ac_results.plot(ax=ax, color='#1f77b4', linewidth=1.0, alpha=0.8)
    ax.set_title("PV System Power Output (2022 - 2024) - ModelChain")
    ax.set_xlabel("Date")
    ax.set_ylabel("Power (W)")
    ax.grid(True, linestyle='--', alpha=0.7)
    fig.tight_layout()
    fig.savefig("results/pv_power_annual_modelchain.png")
    # Removed plt.close(fig)

    # 2. Monthly energy production with cumulative yield
    monthly_sum = ac_results.resample('ME').sum() / 1000  # Convert to kWh
    cumulative_sum = monthly_sum.cumsum()

    fig, ax1 = plt.subplots()
    bars = ax1.bar(monthly_sum.index, monthly_sum.values, width=20, color='#1f77b4', alpha=0.8,
                   label='Monthly Production')
    for bar, value in zip(bars, monthly_sum.values):
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5,
                 f'{value:.0f}', ha='center', va='bottom', fontsize=10, fontweight='bold')

    ax1.set_xlabel('Month')
    ax1.set_ylabel('Monthly Energy Production (kWh)')
    ax1.set_xticks(monthly_sum.index)
    ax1.set_xticklabels([date.strftime('%b') for date in monthly_sum.index], rotation=45)

    ax2 = ax1.twinx()
    ax2.plot(cumulative_sum.index, cumulative_sum.values, 'r-o', linewidth=2.5,
             label='Cumulative Production')
    ax2.set_ylabel('Cumulative Energy Production (kWh)')

    fig.suptitle('Monthly and Cumulative Energy Production (ModelChain)', y=0.98)

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')

    fig.tight_layout()
    fig.savefig('results/monthly_energy_production_modelchain.png')
    # Removed plt.close(fig)

    # 3. Irradiance vs. Power correlation plot
    statistics = {}
    if poa_data is not None and isinstance(poa_data, pd.DataFrame) and 'poa_global' in poa_data.columns:
        try:
            common_idx = ac_results.index.intersection(poa_data.index)

            if len(common_idx) > 0:
                irr_power = pd.DataFrame({
                    'power': ac_results.loc[common_idx].values,
                    'irradiance': poa_data.loc[common_idx, 'poa_global'].values,
                    'month': pd.DatetimeIndex(common_idx).month
                })

                irr_power = irr_power[irr_power['irradiance'] > 50]

                if len(irr_power) > 0:
                    fig, ax = plt.subplots()

                    cmap = plt.cm.viridis
                    scatter = ax.scatter(
                        irr_power['irradiance'],
                        irr_power['power'],
                        c=irr_power['month'],
                        cmap=cmap,
                        alpha=0.6,
                        s=20
                    )

                    try:
                        slope, intercept, r_value, p_value, std_err = stats.linregress(
                            irr_power['irradiance'],
                            irr_power['power']
                        )

                        x = np.array([irr_power['irradiance'].min(), irr_power['irradiance'].max()])
                        y = intercept + slope * x

                        ax.plot(x, y, 'r-', linewidth=2,
                                label=f'Fit: y = {slope:.2f}x + {intercept:.1f}\nR² = {r_value**2:.3f}')
                    except:
                        pass

                    cbar = fig.colorbar(scatter, ax=ax, ticks=range(1,13))
                    cbar.set_label('Month')
                    cbar.ax.set_yticklabels(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                                             'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])

                    ax.set_xlabel('Global POA Irradiance (W/m²)')
                    ax.set_ylabel('AC Power (W)')
                    ax.set_title('Correlation: Solar Irradiance vs. PV Power Output (ModelChain)')
                    ax.grid(True, linestyle='--', alpha=0.7)
                    ax.legend(loc='upper left')

                    fig.tight_layout()
                    fig.savefig('results/irradiance_power_correlation_modelchain.png')
                    # Removed plt.close(fig)
        except Exception as e:
            print(f"Error creating irradiance correlation plot: {e}")
            plt.close('all')

    # 4. Comparison between predicted and actual PV power
    try:
        actual_data = pd.read_csv('data/raw_data_evt_act/merge_resampled.csv', sep=';', decimal=',')

        actual_data['Timestamp'] = pd.to_datetime(actual_data['Timestamp'])
        actual_data.set_index('Timestamp', inplace=True)

        if ac_results.index.tz is not None:
            ac_results.index = ac_results.index.tz_localize(None)
        if actual_data.index.tz is not None:
            actual_data.index = actual_data.index.tz_localize(None)

        start_date = max(ac_results.index.min(), actual_data.index.min())
        end_date = min(ac_results.index.max(), actual_data.index.max())

        predicted_filtered = ac_results[start_date:end_date]
        actual_filtered = actual_data.loc[start_date:end_date, 'PV_Power_W']

        if len(predicted_filtered) != len(actual_filtered):
            actual_filtered = actual_filtered.resample('h').mean()
            predicted_filtered = predicted_filtered.resample('h').mean()

        common_idx = predicted_filtered.index.intersection(actual_filtered.index)

        if len(common_idx) > 0:
            statistics = calculate_statistics(
                predicted_filtered.loc[common_idx],
                actual_filtered.loc[common_idx].values
            )

            fig, ax = plt.subplots(figsize=(12, 6))

            ax.plot(predicted_filtered.loc[common_idx],
                    label='ModelChain Predicted Power', color='blue', linewidth=1.5, alpha=0.8)
            ax.plot(actual_filtered.loc[common_idx],
                    label='Actual Power', color='red', linewidth=1.5, alpha=0.8)

            ax.set_title("Comparison of ModelChain and Actual PV Power Output")
            ax.set_xlabel("Date")
            ax.set_ylabel("Power (W)")
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.legend(loc='upper left')

            fig.tight_layout()
            fig.savefig('results/pv_power_comparison_modelchain.png')
            # Removed plt.close(fig)

            fig, ax = plt.subplots()
            ax.scatter(predicted_filtered.loc[common_idx], actual_filtered.loc[common_idx],
                       alpha=0.5, s=10, label='Model vs Actual')

            max_val = max(
                predicted_filtered.loc[common_idx].max(),
                actual_filtered.loc[common_idx].max()
            )
            ax.plot([0, max_val], [0, max_val], 'r--', label='Perfect Prediction')

            try:
                slope, intercept, r_value, _, _ = stats.linregress(
                    predicted_filtered.loc[common_idx],
                    actual_filtered.loc[common_idx]
                )

                ax.plot([0, max_val], [intercept, intercept + slope * max_val], 'g-',
                        label=f'Fit: y = {slope:.2f}x + {intercept:.1f}\nR² = {r_value**2:.3f}')
            except Exception as e:
                print(f"Error calculating regression: {e}")

            ax.set_xlabel("ModelChain Power (W)")
            ax.set_ylabel("Actual Power (W)")
            ax.set_title("Correlation: ModelChain vs. Actual Power Output")
            ax.grid(True)

            fig.tight_layout()
            fig.savefig('results/pv_power_correlation_model_actual_modelchain.png')
            # Removed plt.close(fig)

    except Exception as e:
        print(f"Error creating comparison plots: {e}")
        plt.close('all') # Keep this here to close figures in case of an error

    return statistics


def compare_approaches() -> None:
    """
    Run both the manual calculation and ModelChain approach and compare results.
    """
    # Import the manual calculation approach
    import pvsimmodule

    # Run both approaches
    print("Running manual calculation approach...")
    ac_manual, poa_manual = pvsimmodule.calculate_power_output(PARAMS)

    print("Running ModelChain approach...")
    ac_modelchain, poa_mc = calculate_power_modelchain(PARAMS)

    # Ensure index alignment
    common_idx = ac_manual.index.intersection(ac_modelchain.index)

    # Calculate differences
    diff = (ac_modelchain.loc[common_idx] - ac_manual.loc[common_idx]).abs()

    # Print comparison statistics
    print("\n" + "="*50)
    print("COMPARISON: MANUAL vs MODEL CHAIN APPROACH")
    print("="*50)
    print(f"Mean absolute difference: {diff.mean():.2f} W")
    print(f"Maximum absolute difference: {diff.max():.2f} W")
    print(f"Relative mean difference: {(diff.mean() / ac_manual.loc[common_idx].mean() * 100):.4f}%")

    # Plot comparison
    fig, ax = plt.subplots(figsize=(10, 6))

    # Plot first week for clarity
    first_week = common_idx[0:168]  # First week (7 days * 24 hours)

    ax.plot(ac_manual.loc[first_week],
            label='Manual Calculation', color='blue', linewidth=1.5)
    ax.plot(ac_modelchain.loc[first_week],
            label='ModelChain', color='red', linewidth=1.5, linestyle='--')

    ax.set_title("Comparison of Manual Calculation vs ModelChain")
    ax.set_xlabel("Date")
    ax.set_ylabel("Power (W)")
    ax.grid(True, linestyle='--', alpha=0.7)
    ax.legend(loc='upper left')

    fig.tight_layout()
    fig.savefig('results/manual_vs_modelchain_comparison.png')
    # Removed plt.close(fig)

    # Create scatter plot
    fig, ax = plt.subplots()
    ax.scatter(ac_manual.loc[common_idx], ac_modelchain.loc[common_idx],
               alpha=0.5, s=5)

    max_val = max(
        ac_manual.loc[common_idx].max(),
        ac_modelchain.loc[common_idx].max()
    )
    ax.plot([0, max_val], [0, max_val], 'r--')

    ax.set_xlabel("Manual Calculation Power (W)")
    ax.set_ylabel("ModelChain Power (W)")
    ax.set_title("Correlation: Manual vs ModelChain Power Output")
    ax.grid(True)

    fig.tight_layout()
    fig.savefig('results/manual_vs_modelchain_correlation.png')
    # Removed plt.close(fig)


# The pvwatts_losses function seems to be a custom implementation or intended as a method
# of the ModelChain class. It is not used in the main flow of the provided script
# and appears to be incomplete or misplaced as a standalone function.
# I will leave it as is, but note that it's not integrated into the main simulation logic.
# If it was intended to replace the built-in pvwatts losses model, it would need
# to be assigned to mc.losses_model and potentially adapted to work with the ModelChain structure.
def pvwatts_losses(self):
    """Apply PVWatts system losses.

    The PVWatts loss model accounts for various system losses such as soiling,
    shading, snow, mismatch, wiring, connections, light-induced degradation (LID),
    nameplate rating, age, and availability.

    Returns
    -------
    self

    Notes
    -----
    Applies the calculated losses factor to the DC output power.
    """
    # Get loss factor as a decimal (1.0 means no losses)
    self.results.losses = (100 - self.system.pvwatts_losses()) / 100.

    # Apply the loss factor to DC power
    if isinstance(self.results.dc, tuple):
        for i, dc in enumerate(self.results.dc):
            if isinstance(dc, pd.DataFrame):
                # For DataFrame format (single diode models)
                dc['p_mp'] *= self.results.losses
            else:
                # For Series format (pvwatts_dc model)
                self.results.dc[i] *= self.results.losses
    else:
        if isinstance(self.results.dc, pd.DataFrame):
            # For DataFrame format (single diode models)
            dc['p_mp'] *= self.results.losses
        else:
            # For Series format (pvwatts_dc model)
            self.results.dc *= self.results.losses

    return self


def main() -> int:
    """
    Main function to run the ModelChain simulation.

    Returns:
        int: Exit status code
    """
    print("Running PV simulation using ModelChain...")

    # Run the ModelChain calculation
    ac_results, poa_data = calculate_power_modelchain(PARAMS)

    # Plot the results
    statistics = plot_results(ac_results, poa_data)

    # Format the Excel file
    # Ensure the Excel file exists before trying to format it
    if os.path.exists('results_modelchain.xlsx'):
        execute_vba_actions('results_modelchain.xlsx')
    else:
        print("Warning: results_modelchain.xlsx not found. Skipping Excel formatting.")


    # Print statistical comparison metrics if available
    if statistics:
        print("\n" + "="*50)
        print("STATISTICAL COMPARISON: MODELCHAIN vs ACTUAL PV POWER")
        print("="*50)
        print(f"RMSE (Root Mean Square Error): {statistics['rmse']:.2f} W")
        print(f"R² (Coefficient of determination): {statistics['r2']:.4f}")
        print(f"MAPE (Mean Absolute Percentage Error): {statistics['mape']:.2f}%")
        print(f"SMAPE (Symmetric Mean Absolute Percentage Error): {statistics['smape']:.2f}%")
        print("="*50 + "\n")

    # Compare with manual approach (uncomment to run comparison)
    # compare_approaches()

    # Keep plots open after execution
    plt.show()

    return 0


if __name__ == '__main__':
    sys.exit(main())
