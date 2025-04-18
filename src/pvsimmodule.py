"""
This script calculates the power output of a photovoltaic system.

The script uses the PVLib library to calculate the power output of a photovoltaic system based on provided parameters.
It also processes and saves the results in an Excel file and generates plots for visualization.

Author: Michael Grün
Email: michaelgruen@hotmail.com
Version: 1.0
Date: 2024-10-13
"""

import os
from typing import Tuple

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import pvlib
from openpyxl.utils import get_column_letter
from pvlib.location import Location
from pvlib.pvsystem import PVSystem
import poadata
from datetime import datetime, timedelta
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy import stats


# Define parameters
PARAMS = {
    'start': '2022-01-01 01:10',
    'end': '2024-01-01 00:10',
    'latitude': 47.3877,
    'longitude': 15.0941,
    'tilt': 30,
    'azimuth': 149.716,
    'celltype': 'polycrystalline',
    'pdc0': 240,
    'v_mp': 29.87,
    'i_mp': 8.04,
    'v_oc': 37.33,
    'i_sc': 8.78,
    'alpha_sc': 0.0041,
    'beta_voc': -0.114,
    'gamma_pdc': -0.405,
    'cells_in_series': 69,
    'temp_ref': 25
}

# Define helper functions
def get_date_attributes(date: datetime) -> Tuple[int, int, int]:
    """
    Return the weekday, hour, and month for a given date.

    Args:
        date (datetime): The date for which attributes are to be extracted.

    Returns:
        Tuple[int, int, int]: A tuple containing the weekday (1-7), hour (0-23), and month (1-12).
    """
    return date.weekday() + 1, date.hour, date.month



def execute_vba_actions(file_name: str) -> None:
    """
    Execute various VBA-like actions on an Excel file to format and adjust column widths.
    Also, copy columns B:J from 'POA Data' to 'Model Chain Results' starting at column H.

    Args:
        file_name (str): The name of the Excel file to be processed.

    Returns:
        None
    """
    workbook = load_workbook(file_name)
    model_chain_results = workbook['Model Chain Results']
    poa_data = workbook['POA Data']

    # Autofit all columns in 'Model Chain Results' and 'POA Data'
    for worksheet in [model_chain_results, poa_data]:
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
            adjusted_width = (max_length + 2) * 1.2
            column_letter = get_column_letter(column[0].column)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Copy columns B:J from 'POA Data' to 'Model Chain Results' starting at column H
    for col_idx, source_col in enumerate(range(2, 11), start=8):  # B:J -> H onwards
        for row_idx, cell in enumerate(poa_data.iter_rows(min_row=1, min_col=source_col, max_col=source_col), start=1):
            model_chain_results.cell(row=row_idx, column=col_idx, value=cell[0].value)

    # Save the modified workbook
    workbook.save(file_name)



def calculate_power_output(params: dict) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Calculate the power output of the photovoltaic system based on provided parameters.

    Args:
        params (dict): A dictionary containing various parameters for the PV system and simulation.

    Returns:
        Tuple[pd.DataFrame, pd.DataFrame]: A tuple containing the AC power results and POA data for 2020.
    """
    location = Location(latitude=params['latitude'], longitude=params['longitude'],
                        tz='Europe/Vienna', altitude=547.6, name='EVT')

    poa_data = poadata.get_pvgis_data(params['latitude'], params['longitude'], 2022, 2023, params['tilt'], params['azimuth'])
    poa_data.to_csv("poa_data_Leoben_EVT_io.csv")
    poa_data = pd.read_csv('poa_data_Leoben_EVT_io.csv', index_col=0)
    poa_data.index = pd.date_range(start='2022-01-01 01:10', periods=len(poa_data.index), freq="h")
    poa_data = poa_data[params['start']:params['end']]

    solarpos = location.get_solarposition(times=pd.date_range(params['start'], end=params['end'], freq="h"))
    aoi = pvlib.irradiance.aoi(params['tilt'], params['azimuth'], solarpos.apparent_zenith, solarpos.azimuth)
    iam = pvlib.iam.ashrae(aoi)
    effective_irradiance = poa_data["poa_direct"] * iam + poa_data["poa_diffuse"]
    temp_cell = pvlib.temperature.faiman(poa_data["poa_global"], poa_data["temp_air"], poa_data["wind_speed"])

    I_L_ref, I_o_ref, R_s, R_sh_ref, a_ref, Adjust = pvlib.ivtools.sdm.fit_cec_sam(
        celltype=params['celltype'], v_mp=params['v_mp'], i_mp=params['i_mp'],
        v_oc=params['v_oc'], i_sc=params['i_sc'], alpha_sc=params['alpha_sc'],
        beta_voc=params['beta_voc'], gamma_pmp=params['gamma_pdc'], cells_in_series=params['cells_in_series']
    )

    cec_params = pvlib.pvsystem.calcparams_cec(effective_irradiance, temp_cell, params['alpha_sc'],
                                               a_ref, I_L_ref, I_o_ref, R_sh_ref, R_s, Adjust)

    mpp = pvlib.pvsystem.max_power_point(*cec_params, method="newton")
    system = PVSystem(modules_per_string=14, strings_per_inverter=3)
    dc_scaled = system.scale_voltage_current_power(mpp)

    cec_inverters = pvlib.pvsystem.retrieve_sam('CECInverter')
    inverter = cec_inverters['Advanced_Energy_Industries__AE_3TL_23_10_08__480V_']
    ac_results = pvlib.inverter.sandia(v_dc=dc_scaled.v_mp, p_dc=dc_scaled.p_mp, inverter=inverter)

    results_df = pd.concat([ac_results, dc_scaled.i_mp, dc_scaled.v_mp, dc_scaled.p_mp, temp_cell], axis=1)
    results_df.columns = ['AC Power', 'DC scaled I_mp', 'DC scaled V_mp', 'DC scaled P_mp', 'Cell Temperature']

    # Save as Excel
    with pd.ExcelWriter("results.xlsx") as writer:
        results_df.to_excel(writer, sheet_name='Model Chain Results')
        poa_data.to_excel(writer, sheet_name='POA Data')

    # Merge datasets and save as CSV
    merged_df = pd.concat([results_df, poa_data], axis=1)
    merged_df.to_csv("merged_results.csv")

    return ac_results, poa_data


def calculate_statistics(predicted: pd.Series, actual: pd.Series) -> dict:
    """
    Calculate various statistical metrics to compare predicted and actual values.
    
    Args:
        predicted (pd.Series): Predicted values
        actual (pd.Series): Actual observed values
        
    Returns:
        dict: Dictionary containing various statistical metrics
    """
    import numpy as np
    
    # Remove any NaN values
    mask = ~(np.isnan(predicted) | np.isnan(actual))
    predicted_clean = predicted[mask]
    actual_clean = actual[mask]
    
    if len(predicted_clean) == 0:
        return {
            'rmse': np.nan,
            'r2': np.nan,
            'mape': np.nan,
            'smape': np.nan
        }
    
    # Calculate RMSE (Root Mean Square Error)
    rmse = np.sqrt(np.mean((predicted_clean - actual_clean) ** 2))
    
    # Calculate R² (Coefficient of determination)
    r2 = stats.pearsonr(predicted_clean, actual_clean)[0] ** 2
    
    # Calculate MAPE (Mean Absolute Percentage Error)
    # Avoid division by zero by excluding zeros in actual values
    non_zero = actual_clean != 0
    mape = np.mean(np.abs((actual_clean[non_zero] - predicted_clean[non_zero]) / actual_clean[non_zero])) * 100
    
    # Calculate SMAPE (Symmetric Mean Absolute Percentage Error)
    # Avoid division by zero
    denominator = np.abs(actual_clean) + np.abs(predicted_clean)
    non_zero_denom = denominator != 0
    smape = np.mean(2.0 * np.abs(predicted_clean[non_zero_denom] - actual_clean[non_zero_denom]) / 
                   denominator[non_zero_denom]) * 100
    
    return {
        'rmse': rmse,
        'r2': r2,
        'mape': mape,
        'smape': smape
    }


def plot_results(ac_results: pd.DataFrame, poa_data: pd.DataFrame = None) -> dict:
    """
    Create high-quality plots.
    
    1. Annual time series of PV power output
    2. Monthly energy production with cumulative yield
    3. PV power vs. irradiance correlation (if POA data is provided)
    4. Comparison between predicted and actual PV power output
    
    Args:
        ac_results (pd.DataFrame): The DataFrame containing the AC power results.
        poa_data (pd.DataFrame, optional): The DataFrame containing POA irradiance data.
    
    Returns:
        dict: Dictionary containing statistical metrics for model performance,
              or empty dict if no comparison was made
    """
    # Set publication-quality plot style
    try:
        plt.style.use('seaborn-v0_8-whitegrid')  # Updated style name
    except:
        plt.style.use('seaborn-whitegrid')  # Fallback for older matplotlib
        
    plt.rcParams.update({
        'font.family': 'serif',
        'font.size': 12,
        'axes.labelsize': 14,
        'axes.titlesize': 16,
        'xtick.labelsize': 12,
        'ytick.labelsize': 12,
        'legend.fontsize': 12,
        'figure.titlesize': 18,
        'figure.figsize': (10, 6),
        'figure.dpi': 300,
        'savefig.dpi': 300,
        'savefig.bbox': 'tight',
        'savefig.pad_inches': 0.1
    })
    
    # Create results directory if it doesn't exist
    os.makedirs('results', exist_ok=True)
    
    
    
    # 1. Annual time series (without seasonal highlighting)
    fig, ax = plt.subplots()
    
    # Plot hourly data directly (no resampling)
    ac_results.plot(ax=ax, color='#1f77b4', linewidth=1.0, alpha=0.8)
    
    # Format plot
    ax.set_title("PV System Power Output (2022 - 2024)")
    ax.set_xlabel("Date")
    ax.set_ylabel("Power (W)")
    ax.grid(True, linestyle='--', alpha=0.7)
    
    fig.tight_layout()
    fig.savefig("results/pv_power_annual.png")
    plt.close(fig)
    
    # 2. Monthly energy production with cumulative yield
    monthly_sum = ac_results.resample('ME').sum() / 1000  # Convert to kWh
    cumulative_sum = monthly_sum.cumsum()
    
    fig, ax1 = plt.subplots()
    
    # Bar chart for monthly production
    bars = ax1.bar(monthly_sum.index, monthly_sum.values, width=20, color='#1f77b4', alpha=0.8,
                  label='Monthly Production')
    
    # Add value labels on top of bars
    for bar, value in zip(bars, monthly_sum.values):
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5,
                f'{value:.0f}', ha='center', va='bottom', fontsize=10, fontweight='bold')
    
    # Format x-axis with month names
    ax1.set_xlabel('Month')
    ax1.set_ylabel('Monthly Energy Production (kWh)')
    ax1.set_xticks(monthly_sum.index)
    ax1.set_xticklabels([date.strftime('%b') for date in monthly_sum.index], rotation=45)
    
    # Secondary y-axis for cumulative production
    ax2 = ax1.twinx()
    ax2.plot(cumulative_sum.index, cumulative_sum.values, 'r-o', linewidth=2.5, 
            label='Cumulative Production')
    ax2.set_ylabel('Cumulative Energy Production (kWh)')
    
    # Add title
    fig.suptitle('Monthly and Cumulative Energy Production (2020)', y=0.98)
    
    # Combine legends from both axes
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
    
    fig.tight_layout()
    fig.savefig('results/monthly_energy_production.png')
    plt.close(fig)
    
    # 3. Irradiance vs. Power correlation plot (if POA data is available)
    if poa_data is not None and isinstance(poa_data, pd.DataFrame) and 'poa_global' in poa_data.columns:
        try:
            # Get common timestamps
            common_idx = ac_results.index.intersection(poa_data.index)
            
            if len(common_idx) > 0:
                # Create a dataframe with both measurements
                irr_power = pd.DataFrame({
                    'power': ac_results.loc[common_idx].values.flatten(),
                    'irradiance': poa_data.loc[common_idx, 'poa_global'].values,
                    'month': pd.DatetimeIndex(common_idx).month
                })
                
                # Filter for daylight hours (irradiance > 50 W/m²)
                irr_power = irr_power[irr_power['irradiance'] > 50]
                
                if len(irr_power) > 0:
                    fig, ax = plt.subplots()
                    
                    # Define season colors based on month
                    cmap = plt.cm.viridis
                    scatter = ax.scatter(
                        irr_power['irradiance'], 
                        irr_power['power'],
                        c=irr_power['month'], 
                        cmap=cmap,
                        alpha=0.6, 
                        s=20
                    )
                    
                    # Add linear regression line
                    try:
                        # Calculate linear regression
                        slope, intercept, r_value, p_value, std_err = stats.linregress(
                            irr_power['irradiance'], 
                            irr_power['power']
                        )
                        
                        # Create line data
                        x = np.array([irr_power['irradiance'].min(), irr_power['irradiance'].max()])
                        y = intercept + slope * x
                        
                        # Plot the line
                        ax.plot(x, y, 'r-', linewidth=2, 
                               label=f'Fit: y = {slope:.2f}x + {intercept:.1f}\nR² = {r_value**2:.3f}')
                    except:
                        pass  # Skip if regression fails
                    
                    # Add colorbar showing months
                    cbar = fig.colorbar(scatter, ax=ax, ticks=range(1,13))
                    cbar.set_label('Month')
                    cbar.ax.set_yticklabels(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                           'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])
                    
                    # Format plot
                    ax.set_xlabel('Global POA Irradiance (W/m²)')
                    ax.set_ylabel('AC Power (W)')
                    ax.set_title('Correlation: Solar Irradiance vs. PV Power Output')
                    ax.grid(True, linestyle='--', alpha=0.7)
                    ax.legend(loc='upper left')
                    
                    fig.tight_layout()
                    fig.savefig('results/irradiance_power_correlation.png')
                    plt.close(fig)
        except Exception as e:
            print(f"Error creating irradiance correlation plot: {e}")
            plt.close('all')
    
    # 4. Comparison between predicted and actual PV power
    statistics = {}
    try:
        # Load actual PV power data from the resampled CSV file
        actual_data = pd.read_csv('data/raw_data_evt_act/merge_resampled.csv', sep=';', decimal=',')
        
        # Convert timestamp to datetime and set as index
        actual_data['Timestamp'] = pd.to_datetime(actual_data['Timestamp'])
        actual_data.set_index('Timestamp', inplace=True)
        
        # Normalize timezone information across all datasets
        # Convert all timestamps to timezone-naive for consistent comparison
        if ac_results.index.tz is not None:
            ac_results.index = ac_results.index.tz_localize(None)
        if actual_data.index.tz is not None:
            actual_data.index = actual_data.index.tz_localize(None)
        
        # Get overlapping time period for both datasets
        start_date = max(ac_results.index.min(), actual_data.index.min())
        end_date = min(ac_results.index.max(), actual_data.index.max())
        
        # Filter datasets to the common time range
        predicted_filtered = ac_results[start_date:end_date]
        actual_filtered = actual_data.loc[start_date:end_date, 'PV_Power_W']
        
        # Ensure timestamps align by resampling if necessary
        if len(predicted_filtered) != len(actual_filtered):
            actual_filtered = actual_filtered.resample('h').mean()
            predicted_filtered = predicted_filtered.resample('h').mean()
        
        # Find common timestamps across both datasets
        common_idx = predicted_filtered.index.intersection(actual_filtered.index)
        
        if len(common_idx) > 0:
            # Calculate statistics for the comparison
            statistics = calculate_statistics(
                predicted_filtered.loc[common_idx].values.flatten(),
                actual_filtered.loc[common_idx].values
            )
            
            # Create comparison plot
            fig, ax = plt.subplots(figsize=(12, 6))
            
            # Plot both datasets
            ax.plot(predicted_filtered.loc[common_idx], 
                   label='Model Predicted Power', color='blue', linewidth=1.5, alpha=0.8)
            ax.plot(actual_filtered.loc[common_idx], 
                   label='Actual Power', color='red', linewidth=1.5, alpha=0.8)
            
            # Format plot
            ax.set_title("Comparison of Model and Actual PV Power Output")
            ax.set_xlabel("Date")
            ax.set_ylabel("Power (W)")
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.legend(loc='upper left')
            
            fig.tight_layout()
            fig.savefig('results/pv_power_comparison.png')
            plt.close(fig)
            
            # Create scatter plots for model vs actual
            fig, ax = plt.subplots()
            ax.scatter(predicted_filtered.loc[common_idx], actual_filtered.loc[common_idx],
                      alpha=0.5, s=10, label='Model vs Actual')
            
            # Add reference line (perfect prediction)
            max_val = max(
                predicted_filtered.loc[common_idx].max(), 
                actual_filtered.loc[common_idx].max()
            )
            ax.plot([0, max_val], [0, max_val], 'r--', label='Perfect Prediction')
            
            # Add linear regression
            try:
                slope, intercept, r_value, _, _ = stats.linregress(
                    predicted_filtered.loc[common_idx], 
                    actual_filtered.loc[common_idx]
                )
                
                ax.plot([0, max_val], [intercept, intercept + slope * max_val], 'g-',
                      label=f'Fit: y = {slope:.2f}x + {intercept:.1f}\nR² = {r_value**2:.3f}')
            except Exception as e:
                print(f"Error calculating regression: {e}")
            
            ax.set_xlabel("Model Power (W)")
            ax.set_ylabel("Actual Power (W)")
            ax.set_title("Correlation: Model vs. Actual Power Output")
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.legend(loc='upper left')
            
            fig.tight_layout()
            fig.savefig('results/pv_power_correlation_model_actual.png')
            plt.close(fig)
            
    except Exception as e:
        print(f"Error creating comparison plots: {e}")
        plt.close('all')
    
    return statistics


def main() -> int:
    """
    Main function to run the power output calculation, plotting, and Excel file processing.

    Returns:
        int: The exit status code.
    """
    ac_results, poa_data = calculate_power_output(PARAMS)
    statistics = plot_results(ac_results, poa_data)
    execute_vba_actions('results.xlsx')
    
    # Print statistical comparison metrics if available
    if statistics:
        print("\n" + "="*50)
        print("STATISTICAL COMPARISON: PREDICTED vs ACTUAL PV POWER")
        print("="*50)
        print(f"RMSE (Root Mean Square Error): {statistics['rmse']:.2f} W")
        print(f"R² (Coefficient of determination): {statistics['r2']:.4f}")
        print(f"MAPE (Mean Absolute Percentage Error): {statistics['mape']:.2f}%")
        print(f"SMAPE (Symmetric Mean Absolute Percentage Error): {statistics['smape']:.2f}%")
        print("="*50 + "\n")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
