"""
This script calculates the power output of a photovoltaic system.

The script uses the PVLib library to calculate the power output of a photovoltaic system based on provided parameters.
It also processes and saves the results in an Excel file and generates plots for visualization.

Author: Michael Grün
Email: michaelgruen@hotmail.com
Version: 1.0
Date: 2024-10-13
"""

import os
from typing import Tuple

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import pvlib
from openpyxl.utils import get_column_letter
from pvlib.location import Location
from pvlib.pvsystem import PVSystem
import poadata
from datetime import datetime, timedelta
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Define parameters
PARAMS = {
    'start': '2020-01-01 00:00',
    'end': '2020-12-31 23:00',
    'latitude': 47.3877,
    'longitude': 15.0941,
    'tilt': 30,
    'azimuth': 149.716,
    'celltype': 'polycrystalline',
    'pdc0': 240,
    'v_mp': 29.87,
    'i_mp': 8.04,
    'v_oc': 37.33,
    'i_sc': 8.78,
    'alpha_sc': 0.0041,
    'beta_voc': -0.114,
    'gamma_pdc': -0.405,
    'cells_in_series': 69,
    'temp_ref': 25
}

# Define helper functions
def get_date_attributes(date: datetime) -> Tuple[int, int, int]:
    """
    Return the weekday, hour, and month for a given date.

    Args:
        date (datetime): The date for which attributes are to be extracted.

    Returns:
        Tuple[int, int, int]: A tuple containing the weekday (1-7), hour (0-23), and month (1-12).
    """
    return date.weekday() + 1, date.hour, date.month



def execute_vba_actions(file_name: str) -> None:
    """
    Execute various VBA-like actions on an Excel file to format and adjust column widths.
    Also, copy columns B:J from 'POA Data' to 'Model Chain Results' starting at column H.

    Args:
        file_name (str): The name of the Excel file to be processed.

    Returns:
        None
    """
    workbook = load_workbook(file_name)
    model_chain_results = workbook['Model Chain Results']
    poa_data = workbook['POA Data']

    # Autofit all columns in 'Model Chain Results' and 'POA Data'
    for worksheet in [model_chain_results, poa_data]:
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
            adjusted_width = (max_length + 2) * 1.2
            column_letter = get_column_letter(column[0].column)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Copy columns B:J from 'POA Data' to 'Model Chain Results' starting at column H
    for col_idx, source_col in enumerate(range(2, 11), start=8):  # B:J -> H onwards
        for row_idx, cell in enumerate(poa_data.iter_rows(min_row=1, min_col=source_col, max_col=source_col), start=1):
            model_chain_results.cell(row=row_idx, column=col_idx, value=cell[0].value)

    # Save the modified workbook
    workbook.save(file_name)



def calculate_power_output(params: dict) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Calculate the power output of the photovoltaic system based on provided parameters.

    Args:
        params (dict): A dictionary containing various parameters for the PV system and simulation.

    Returns:
        Tuple[pd.DataFrame, pd.DataFrame]: A tuple containing the AC power results and POA data for 2020.
    """
    location = Location(latitude=params['latitude'], longitude=params['longitude'],
                        tz='Europe/Vienna', altitude=547.6, name='EVT')

    poa_data_2020 = poadata.get_pvgis_data(params['latitude'], params['longitude'], 2020, 2020, params['tilt'], params['azimuth'])
    poa_data_2020.to_csv("poa_data_2020_Leoben_EVT_io.csv")
    poa_data_2020 = pd.read_csv('poa_data_2020_Leoben_EVT_io.csv', index_col=0)
    poa_data_2020.index = pd.date_range(start='2020-01-01 00:00', periods=len(poa_data_2020.index), freq="h")
    poa_data = poa_data_2020[params['start']:params['end']]

    solarpos = location.get_solarposition(times=pd.date_range(params['start'], end=params['end'], freq="h"))
    aoi = pvlib.irradiance.aoi(params['tilt'], params['azimuth'], solarpos.apparent_zenith, solarpos.azimuth)
    iam = pvlib.iam.ashrae(aoi)
    effective_irradiance = poa_data["poa_direct"] + iam + poa_data["poa_diffuse"]
    temp_cell = pvlib.temperature.faiman(poa_data["poa_global"], poa_data["temp_air"], poa_data["wind_speed"])

    I_L_ref, I_o_ref, R_s, R_sh_ref, a_ref, Adjust = pvlib.ivtools.sdm.fit_cec_sam(
        celltype=params['celltype'], v_mp=params['v_mp'], i_mp=params['i_mp'],
        v_oc=params['v_oc'], i_sc=params['i_sc'], alpha_sc=params['alpha_sc'],
        beta_voc=params['beta_voc'], gamma_pmp=params['gamma_pdc'], cells_in_series=params['cells_in_series']
    )

    cec_params = pvlib.pvsystem.calcparams_cec(effective_irradiance, temp_cell, params['alpha_sc'],
                                               a_ref, I_L_ref, I_o_ref, R_sh_ref, R_s, Adjust)

    mpp = pvlib.pvsystem.max_power_point(*cec_params, method="newton")
    system = PVSystem(modules_per_string=23, strings_per_inverter=3)
    dc_scaled = system.scale_voltage_current_power(mpp)

    cec_inverters = pvlib.pvsystem.retrieve_sam('CECInverter')
    inverter = cec_inverters['Advanced_Energy_Industries__AE_3TL_23_10_08__480V_']
    ac_results = pvlib.inverter.sandia(v_dc=dc_scaled.v_mp, p_dc=dc_scaled.p_mp, inverter=inverter)

    results_df = pd.concat([ac_results, dc_scaled.i_mp, dc_scaled.v_mp, dc_scaled.p_mp, temp_cell], axis=1)
    results_df.columns = ['AC Power', 'DC scaled I_mp', 'DC scaled V_mp', 'DC scaled P_mp', 'Cell Temperature']

    # Save as Excel
    with pd.ExcelWriter("results.xlsx") as writer:
        results_df.to_excel(writer, sheet_name='Model Chain Results')
        poa_data_2020.to_excel(writer, sheet_name='POA Data')

    # Merge datasets and save as CSV
    merged_df = pd.concat([results_df, poa_data_2020], axis=1)
    merged_df.to_csv("merged_results.csv")

    return ac_results, poa_data_2020


def plot_results(ac_results: pd.DataFrame) -> None:
    """
    Plot the AC power results over time and as a monthly sum.

    Args:
        ac_results (pd.DataFrame): The DataFrame containing the AC power results.

    Returns:
        None
    """
    ac_results.plot(figsize=(16, 9))
    plt.title("AC Power - PVSystem")
    plt.xlabel("Time")
    plt.ylabel("Energy Yield")
    plt.grid(True)
    plt.savefig("energy_yield_start_to_end.png")
    plt.show()

    monthly_sum = ac_results.resample('M').sum()
    monthly_sum.plot(figsize=(16, 9))
    plt.title("AC Power - PVSystem (Monthly Sum)")
    plt.xlabel("Time")
    plt.ylabel("Energy Yield")
    plt.grid(True)
    plt.savefig("energy_yield_monthly_sum.png")
    plt.show()


def main() -> int:
    """
    Main function to run the power output calculation, plotting, and Excel file processing.

    Returns:
        int: The exit status code.
    """
    ac_results, poa_data_2020 = calculate_power_output(PARAMS)
    plot_results(ac_results)
    execute_vba_actions('results.xlsx')
    return 0


if __name__ == '__main__':
    sys.exit(main())